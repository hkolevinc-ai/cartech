import os
import re
import json
import html
import time
import logging
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Set, Tuple
from urllib.parse import urljoin, urlparse, urlunparse, parse_qsl, urlencode

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_URL = os.getenv("BASE_URL", "https://cartechbg.com/").rstrip("/") + "/"
OUTPUT_FILE = os.getenv("OUTPUT_FILE", "cartechbg_all_products.xlsx")
REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "30"))
REQUEST_SLEEP = float(os.getenv("REQUEST_SLEEP", "0.35"))
MAX_PRODUCTS = int(os.getenv("MAX_PRODUCTS", "0"))  # 0 = no limit
MAX_SHOP_PAGES = int(os.getenv("MAX_SHOP_PAGES", "200"))
MAX_CATEGORY_PAGES = int(os.getenv("MAX_CATEGORY_PAGES", "120"))
MAX_IMAGES = int(os.getenv("MAX_IMAGES", "7"))  # requested columns 5-11

HEADERS = {
    "User-Agent": os.getenv(
        "USER_AGENT",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/125.0 Safari/537.36 CarTechBGScraper/3.0",
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "bg-BG,bg;q=0.9,en;q=0.8",
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("cartechbg-scraper")

session = requests.Session()
session.headers.update(HEADERS)


# Categories from the screenshots / menu supplied by the user.
# The final Excel column "Категория" is the most specific matching label from this list.
# Keep specific categories before broader parent categories so classification prefers subcategories.
CATEGORY_DEFINITIONS: List[Tuple[str, str]] = [
    # Екстериор
    ("Angel Eyes", "https://cartechbg.com/product-category/eksterior/angel-eyes/"),
    ("Вежди за Фарове", "https://cartechbg.com/product-category/eksterior/vejdi-farove/"),
    ("Ветробрани", "https://cartechbg.com/product-category/eksterior/vetrobrani/"),
    ("Дифузори и Добавки за Задни брони", "https://cartechbg.com/product-category/eksterior/difuziori-dobavki-zadna-bronya/"),
    ("Добавки за Прагове", "https://cartechbg.com/product-category/eksterior/dobavki-pragove/"),
    ("Добавки за Предна Броня", "https://cartechbg.com/product-category/eksterior/dobavki-predna-bronya/"),
    ("Добавки за Стъкла", "https://cartechbg.com/product-category/eksterior/dobavki-stukla/"),
    ("Калници", "https://cartechbg.com/product-category/eksterior/kalnici/"),
    ("Капачки за Джанти", "https://cartechbg.com/product-category/eksterior/kapachki-djanti/"),
    ("Капаци за Огледала", "https://cartechbg.com/product-category/eksterior/kapaci-ogledala-batman/"),
    ("Капачки за Резервоар", "https://cartechbg.com/product-category/eksterior/kapachki-rezervoar/"),
    ("Капачки и Решетки за Брони", "https://cartechbg.com/product-category/eksterior/kapachki-reshetki-broni/"),
    ("Накрайници за Ауспуси", "https://cartechbg.com/product-category/eksterior/nakraynici-auspusi/"),
    ("Окачвания", "https://cartechbg.com/product-category/chaste/okachvaniq/"),
    ("Прагове", "https://cartechbg.com/product-category/eksterior/pragove/"),
    ("Решетки и Бъбреци", "https://cartechbg.com/product-category/eksterior/reshetki-bubreci/"),
    ("Спойлери", "https://cartechbg.com/product-category/eksterior/spoyleri/"),
    ("Стъкла за Фарове", "https://cartechbg.com/product-category/eksterior/stukla-farove/"),
    ("Тунинг Брони и Пакети", "https://cartechbg.com/product-category/eksterior/tuning-broni-paketi/"),
    ("Тунинг Мигачи", "https://cartechbg.com/product-category/eksterior/tuning-migachi/"),
    ("Тунинг Фарове и Стопове", "https://cartechbg.com/product-category/eksterior/tuning-farove-stopove/"),
    ("Халогени", "https://cartechbg.com/product-category/eksterior/halogeni/"),
    # Интериор
    ("Волани", "https://cartechbg.com/product-category/interior/volani/"),
    ("Въздуховоди", "https://cartechbg.com/product-category/interior/vuzduhovodi/"),
    ("Дигитални Табла", "https://cartechbg.com/product-category/interior/digitalni-tabla/"),
    ("Дръжки и Панели", "https://cartechbg.com/product-category/interior/vutreshni-drujki-bmw/"),
    ("Интериорни Елементи", "https://cartechbg.com/product-category/interior/butoni-vurtoci/"),
    ("Къпхолдъри", "https://cartechbg.com/product-category/interior/cupholders/"),
    ("Пера за Волан", "https://cartechbg.com/product-category/interior/pera-volan/"),
    ("Приставки за Волан", "https://cartechbg.com/product-category/interior/pristavki-za-volan/"),
    ("Скоростни лостове", "https://cartechbg.com/product-category/interior/skorostni-lostove/"),
    ("Стелки", "https://cartechbg.com/product-category/interior/stelki/"),
    ("Универсални аксесоари", "https://cartechbg.com/product-category/interior/universalni-aksesoari/"),
    # Адаптери / мултимедия
    ("Адаптери", "https://cartechbg.com/product-category/adapteri-carplay-androidauto-tv-yt/"),
    ("Мултимедии и Интерфейси", "https://cartechbg.com/product-category/multimedii-interfeisi/"),
    # Консумативи и аксесоари
    ("Аксесоари", "https://cartechbg.com/product-category/konsumativi-aksesoari-avtomobili/aksesoari/"),
    ("Добавки", "https://cartechbg.com/product-category/konsumativi-aksesoari-avtomobili/dobavki/"),
    ("Крушки", "https://cartechbg.com/product-category/konsumativi-aksesoari-avtomobili/krushki-led-xenon-cartech/"),
    ("Грижа за автомобила", "https://cartechbg.com/product-category/grija-za-avtomobila/"),
    ("Чистачки за коли", "https://cartechbg.com/product-category/konsumativi-aksesoari-avtomobili/chistachki-za-koli/"),
    # Broad fallback labels from the menu
    ("Екстериор", "https://cartechbg.com/product-category/eksterior/"),
    ("Интериор", "https://cartechbg.com/product-category/interior/"),
    ("Консумативи и Аксесоари", "https://cartechbg.com/product-category/konsumativi-aksesoari-avtomobili/"),
]

CATEGORY_PRIORITY = {label: idx for idx, (label, _) in enumerate(CATEGORY_DEFINITIONS)}
CATEGORY_BY_SLUG: Dict[str, str] = {}
CATEGORY_BY_NAME: Dict[str, str] = {}
for label, url in CATEGORY_DEFINITIONS:
    slug = urlparse(url).path.rstrip("/").split("/")[-1]
    CATEGORY_BY_SLUG[slug] = label
    CATEGORY_BY_NAME[re.sub(r"\s+", " ", label).strip().casefold()] = label
# Known spelling/typing variants found in the site/menu.
CATEGORY_BY_NAME.update({
    "дифузьори и добавки за задни брони": "Дифузори и Добавки за Задни брони",
    "kапачки за джанти": "Капачки за Джанти",
})

KEYWORD_FALLBACKS: List[Tuple[str, List[str]]] = [
    ("Angel Eyes", ["angel eyes"]),
    ("Вежди за Фарове", ["вежди", "vejdi"]),
    ("Ветробрани", ["ветробран", "vetrobran"]),
    ("Дифузори и Добавки за Задни брони", ["дифуз", "задна брон", "задни брон"]),
    ("Добавки за Прагове", ["добавки за праг", "добавка за праг"]),
    ("Добавки за Предна Броня", ["предна брон", "добавка за броня"]),
    ("Добавки за Стъкла", ["добавки за стъкла", "стъкла"]),
    ("Калници", ["калник"]),
    ("Капачки за Джанти", ["капачки за джанти", "капачка за джанта"]),
    ("Капаци за Огледала", ["капаци за огледала", "капак за огледало", "огледало"]),
    ("Капачки за Резервоар", ["резервоар"]),
    ("Капачки и Решетки за Брони", ["решетки за брони", "капачки за брони"]),
    ("Накрайници за Ауспуси", ["накрайник", "ауспух"]),
    ("Окачвания", ["окачван", "coilover", "амортисьор"]),
    ("Прагове", ["прагове", "праг"]),
    ("Решетки и Бъбреци", ["решетка", "решетки", "бъбрек", "бъбреци"]),
    ("Спойлери", ["спойлер"]),
    ("Стъкла за Фарове", ["стъкло за фар", "стъкла за фар"]),
    ("Тунинг Брони и Пакети", ["тунинг брон", "m-paket", "m пакет", "пакет"]),
    ("Тунинг Мигачи", ["мигач"]),
    ("Тунинг Фарове и Стопове", ["фар", "стоп"]),
    ("Халогени", ["халоген"]),
    ("Волани", ["волан"]),
    ("Въздуховоди", ["въздуховод"]),
    ("Дигитални Табла", ["дигитално табло", "дигитални табла", "табло"]),
    ("Дръжки и Панели", ["дръжк", "панел"]),
    ("Интериорни Елементи", ["интериор"]),
    ("Къпхолдъри", ["къпхолд", "cupholder"]),
    ("Пера за Волан", ["пера за волан", "пера"]),
    ("Приставки за Волан", ["приставки за волан", "приставка"]),
    ("Скоростни лостове", ["скоростен лост", "скоростни лостове"]),
    ("Стелки", ["стелки", "стелка"]),
    ("Адаптери", ["carplay", "androidauto", "android auto", "адаптер"]),
    ("Мултимедии и Интерфейси", ["мултимедия", "интерфейс"]),
    ("Добавки", ["добавка", "добавки"]),
    ("Крушки", ["крушка", "крушки", "led", "xenon", "ксенон"]),
    ("Грижа за автомобила", ["грижа за автомобила", "препарат", "почист", "шампоан", "полира", "микрофибър"]),
    ("Чистачки за коли", ["чистачки", "чистачка"]),
    ("Аксесоари", ["аксесоар"]),
]


@dataclass
class ProductRow:
    sku: str = ""
    product_name: str = ""
    category: str = ""
    description: str = ""
    price_eur: str = ""
    product_url: str = ""
    images: List[str] = None

    def __post_init__(self):
        if self.images is None:
            self.images = []


def fetch(url: str, accept_xml: bool = False) -> Optional[str]:
    """Fetch URL with retries. Returns text or None."""
    for attempt in range(1, 4):
        try:
            headers = dict(HEADERS)
            if accept_xml:
                headers["Accept"] = "application/xml,text/xml,*/*;q=0.8"
            r = session.get(url, timeout=REQUEST_TIMEOUT, headers=headers)
            if r.status_code in (403, 429, 500, 502, 503, 504):
                raise requests.HTTPError(f"HTTP {r.status_code}")
            if r.status_code == 404:
                return None
            r.raise_for_status()
            return r.text
        except Exception as exc:
            wait = attempt * 2
            log.warning("Fetch failed (%s/%s): %s | %s", attempt, 3, url, exc)
            if attempt < 3:
                time.sleep(wait)
    return None


def normalize_url(url: str) -> str:
    url = str(url or "").strip()
    if not url:
        return ""
    absolute = urljoin(BASE_URL, url).split("#")[0]
    parsed = urlparse(absolute)
    # Drop common tracking params but keep WooCommerce pagination params.
    params = [(k, v) for k, v in parse_qsl(parsed.query, keep_blank_values=True) if not k.lower().startswith("utm_")]
    query = urlencode(params)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", query, ""))


def canonical_product_url(url: str) -> str:
    url = normalize_url(url)
    parsed = urlparse(url)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path.rstrip("/") + "/", "", "", ""))


def is_product_url(url: str) -> bool:
    parsed = urlparse(url)
    if not parsed.netloc.endswith("cartechbg.com"):
        return False
    path = parsed.path.rstrip("/") + "/"
    return "/product/" in path and "/product-category/" not in path


def is_category_url(url: str) -> bool:
    parsed = urlparse(url)
    return parsed.netloc.endswith("cartechbg.com") and "/product-category/" in parsed.path


def category_page_url(base_url: str, page: int) -> str:
    base_url = normalize_url(base_url)
    if page <= 1:
        return base_url
    parsed = urlparse(base_url)
    params = [(k, v) for k, v in parse_qsl(parsed.query, keep_blank_values=True) if k != "product-page"]
    params.append(("product-page", str(page)))
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", urlencode(params), ""))


def parse_xml_locs(xml_text: str) -> List[str]:
    soup = BeautifulSoup(xml_text, "xml")
    return [loc.get_text(strip=True) for loc in soup.find_all("loc") if loc.get_text(strip=True)]


def collect_sitemap_product_urls() -> List[str]:
    """Collect product URLs from common WordPress/WooCommerce sitemap locations."""
    candidates = [
        urljoin(BASE_URL, "sitemap.xml"),
        urljoin(BASE_URL, "sitemap_index.xml"),
        urljoin(BASE_URL, "product-sitemap.xml"),
        urljoin(BASE_URL, "product-sitemap1.xml"),
        urljoin(BASE_URL, "product-sitemap2.xml"),
        urljoin(BASE_URL, "wp-sitemap-posts-product-1.xml"),
    ]

    sitemap_urls: Set[str] = set()
    product_urls: Set[str] = set()

    for sitemap in candidates:
        text = fetch(sitemap, accept_xml=True)
        if not text:
            continue
        locs = parse_xml_locs(text)
        for loc in locs:
            l = loc.lower()
            if is_product_url(loc):
                product_urls.add(canonical_product_url(loc))
            elif any(token in l for token in ["product-sitemap", "wp-sitemap-posts-product"]):
                sitemap_urls.add(normalize_url(loc))

    # Probe product-sitemapN.xml until several misses in a row.
    misses = 0
    for i in range(1, 31):
        suffix = "" if i == 1 else str(i)
        probe = urljoin(BASE_URL, f"product-sitemap{suffix}.xml")
        if probe in sitemap_urls:
            continue
        text = fetch(probe, accept_xml=True)
        if not text:
            misses += 1
            if misses >= 4:
                break
            continue
        misses = 0
        locs = parse_xml_locs(text)
        found_products = [canonical_product_url(loc) for loc in locs if is_product_url(loc)]
        if found_products:
            product_urls.update(found_products)
        else:
            sitemap_urls.add(probe)

    for sitemap in sorted(sitemap_urls):
        text = fetch(sitemap, accept_xml=True)
        if not text:
            continue
        for loc in parse_xml_locs(text):
            if is_product_url(loc):
                product_urls.add(canonical_product_url(loc))

    urls = sorted(product_urls)
    log.info("Product URLs from sitemap: %s", len(urls))
    return urls


def extract_product_links_from_page(html_text: str) -> Set[str]:
    soup = BeautifulSoup(html_text, "lxml")
    urls: Set[str] = set()
    for a in soup.select("a[href]"):
        href = canonical_product_url(a.get("href", ""))
        if is_product_url(href):
            urls.add(href)
    return urls


def max_product_page_from_pagination(html_text: str) -> int:
    soup = BeautifulSoup(html_text, "lxml")
    max_page = 1
    for a in soup.select("a[href]"):
        href = a.get("href", "")
        txt = clean_text(a.get_text(" ", strip=True))
        for candidate in [txt, href]:
            for m in re.finditer(r"product-page[=/](\d+)|/page/(\d+)/", candidate):
                num = m.group(1) or m.group(2)
                if num and num.isdigit():
                    max_page = max(max_page, int(num))
        if txt.isdigit():
            max_page = max(max_page, int(txt))
    return min(max_page, MAX_CATEGORY_PAGES)


def better_category(current: str, new: str) -> str:
    if not new:
        return current
    if not current:
        return new
    return new if CATEGORY_PRIORITY.get(new, 9999) < CATEGORY_PRIORITY.get(current, 9999) else current


def collect_category_product_urls() -> Tuple[List[str], Dict[str, str]]:
    """Crawl the user-provided menu categories and keep the best category hint per product URL."""
    product_urls: Set[str] = set()
    hints: Dict[str, str] = {}

    for category_label, cat_url in CATEGORY_DEFINITIONS:
        first_html = fetch(cat_url)
        if not first_html:
            continue
        first_links = extract_product_links_from_page(first_html)
        for url in first_links:
            product_urls.add(url)
            hints[url] = better_category(hints.get(url, ""), category_label)
        last_page = max_product_page_from_pagination(first_html)
        log.info("Category '%s' first page: %s products, pages=%s", category_label, len(first_links), last_page)

        for page in range(2, last_page + 1):
            page_url = category_page_url(cat_url, page)
            html_text = fetch(page_url)
            if not html_text:
                continue
            links = extract_product_links_from_page(html_text)
            if not links:
                continue
            for url in links:
                product_urls.add(url)
                hints[url] = better_category(hints.get(url, ""), category_label)
            log.info("Category '%s' page %s: %s products", category_label, page, len(links))
            time.sleep(REQUEST_SLEEP)
        time.sleep(REQUEST_SLEEP)

    urls = sorted(product_urls)
    log.info("Product URLs from user menu categories: %s", len(urls))
    return urls, hints


def collect_shop_product_urls() -> List[str]:
    """Fallback crawler through shop pagination."""
    product_urls: Set[str] = set()
    stop_after_empty = 0

    patterns = [
        urljoin(BASE_URL, "shop/"),
        urljoin(BASE_URL, "shop/?product-page={page}"),
        urljoin(BASE_URL, "shop/page/{page}/"),
    ]

    for page in range(1, MAX_SHOP_PAGES + 1):
        page_urls = []
        if page == 1:
            page_urls.append(patterns[0])
        page_urls.append(patterns[1].format(page=page))
        page_urls.append(patterns[2].format(page=page))

        before = len(product_urls)
        for page_url in page_urls:
            html_text = fetch(page_url)
            if not html_text:
                continue
            product_urls.update(extract_product_links_from_page(html_text))
        gained = len(product_urls) - before
        log.info("Shop page %s gained %s product URLs; total %s", page, gained, len(product_urls))
        if gained == 0:
            stop_after_empty += 1
            if stop_after_empty >= 5:
                break
        else:
            stop_after_empty = 0
        time.sleep(REQUEST_SLEEP)

    urls = sorted(product_urls)
    log.info("Product URLs from shop fallback: %s", len(urls))
    return urls


def clean_text(text: str) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def parse_price_eur(text: str) -> str:
    if not text:
        return ""
    text = html.unescape(str(text)).replace("\xa0", " ")
    # Prefer values followed by euro sign or EUR.
    matches = re.findall(r"(\d+[\d\s,.]*)\s*(?:€|EUR|евро)", text, flags=re.I)
    if not matches:
        return ""
    raw = matches[-1].replace(" ", "")
    # Convert European decimal comma to dot when needed.
    if raw.count(",") == 1 and raw.count(".") == 0:
        raw = raw.replace(",", ".")
    # Remove thousands separators safely.
    if raw.count(".") > 1:
        parts = raw.split(".")
        raw = "".join(parts[:-1]) + "." + parts[-1]
    return raw


def extract_json_ld(soup: BeautifulSoup) -> List[dict]:
    items = []
    for script in soup.find_all("script", type=lambda v: v and "ld+json" in v):
        raw = script.string or script.get_text() or ""
        raw = raw.strip()
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except Exception:
            continue
        stack = data if isinstance(data, list) else [data]
        while stack:
            obj = stack.pop(0)
            if isinstance(obj, dict):
                items.append(obj)
                graph = obj.get("@graph")
                if isinstance(graph, list):
                    stack.extend(graph)
            elif isinstance(obj, list):
                stack.extend(obj)
    return items


def first_product_json_ld(soup: BeautifulSoup) -> dict:
    for item in extract_json_ld(soup):
        typ = item.get("@type")
        types = typ if isinstance(typ, list) else [typ]
        if any(str(t).lower() == "product" for t in types):
            return item
    return {}


def unique_keep_order(values: Iterable[str]) -> List[str]:
    seen = set()
    out = []
    for v in values:
        v = normalize_url(str(v))
        if not v or v in seen or v.startswith("data:"):
            continue
        seen.add(v)
        out.append(v)
    return out


def unique_text_keep_order(values: Iterable[str]) -> List[str]:
    seen = set()
    out = []
    for v in values:
        v = clean_text(str(v))
        if not v:
            continue
        key = v.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(v)
    return out


def category_slug_from_url(url: str) -> str:
    parsed = urlparse(normalize_url(url))
    return parsed.path.rstrip("/").split("/")[-1]


def mapped_category_from_name_or_url(name: str = "", url: str = "") -> str:
    name_key = clean_text(name).casefold()
    if name_key in CATEGORY_BY_NAME:
        return CATEGORY_BY_NAME[name_key]
    if url:
        slug = category_slug_from_url(url)
        if slug in CATEGORY_BY_SLUG:
            return CATEGORY_BY_SLUG[slug]
    return ""


def extract_category_candidates(soup: BeautifulSoup, product_json: dict) -> List[Tuple[str, str]]:
    candidates: List[Tuple[str, str]] = []

    for a in soup.select(".posted_in a[href], nav.woocommerce-breadcrumb a[href], .breadcrumb a[href]"):
        text = clean_text(a.get_text(" ", strip=True))
        href = normalize_url(a.get("href", ""))
        if text:
            candidates.append((text, href))

    raw_category = clean_text(str(product_json.get("category", "") or ""))
    if raw_category:
        for part in re.split(r"\s*(?:>|/|»|›|,|\|)\s*", raw_category):
            part = clean_text(part)
            if part:
                candidates.append((part, ""))

    ignored = {"начало", "home", "shop", "магазин", "всички продукти", "uncategorized"}
    filtered = []
    for name, href in candidates:
        if name.casefold() in ignored:
            continue
        filtered.append((name, href))
    return filtered


def choose_category(soup: BeautifulSoup, product_json: dict, category_hint: str, product_name: str, product_url: str) -> str:
    best = category_hint or ""

    for name, href in extract_category_candidates(soup, product_json):
        mapped = mapped_category_from_name_or_url(name, href)
        best = better_category(best, mapped)

    if best:
        return best

    haystack_parts = [product_name, product_url]
    for name, href in extract_category_candidates(soup, product_json):
        haystack_parts.extend([name, href])
    haystack = " ".join(haystack_parts).casefold()
    for label, keywords in KEYWORD_FALLBACKS:
        if any(k.casefold() in haystack for k in keywords):
            return label

    # Last fallback: keep the first actual product category from the page, even if not in the supplied menu.
    for name, _href in extract_category_candidates(soup, product_json):
        if name:
            return name
    return ""


def extract_variations(soup: BeautifulSoup, html_text: str) -> List[dict]:
    variations: List[dict] = []

    for form in soup.select("form.variations_form[data-product_variations]"):
        raw = form.get("data-product_variations") or ""
        raw = html.unescape(raw).strip()
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except Exception:
            data = None
        if isinstance(data, list):
            variations.extend([item for item in data if isinstance(item, dict)])

    # Fallback for themes that print variation JSON inside scripts.
    if not variations:
        for m in re.finditer(r'"variation_id"\s*:\s*(\d+).*?"sku"\s*:\s*"([^"\\]*(?:\\.[^"\\]*)*)"', html_text, re.S):
            try:
                sku = json.loads('"' + m.group(2) + '"')
            except Exception:
                sku = m.group(2)
            variations.append({"variation_id": m.group(1), "sku": sku})

    seen = set()
    unique = []
    for item in variations:
        sku = clean_text(item.get("sku", ""))
        vid = clean_text(item.get("variation_id", ""))
        key = sku or vid or json.dumps(item, sort_keys=True, ensure_ascii=False)
        if key and key not in seen:
            seen.add(key)
            unique.append(item)
    return unique


def extract_sku(soup: BeautifulSoup, product_json: dict, product_url: str) -> str:
    sku = clean_text(str(product_json.get("sku", "") or ""))
    if not sku:
        el = soup.select_one(".sku")
        sku = clean_text(el.get_text(" ", strip=True)) if el else ""
    if not sku:
        sku_meta = soup.select_one('[itemprop="sku"], meta[property="product:retailer_item_id"]')
        if sku_meta:
            sku = clean_text(sku_meta.get("content") or sku_meta.get_text(" ", strip=True))
    if not sku:
        # Stable fallback from Woo product id or URL slug, to keep the SKU column unique/non-empty.
        product_id = ""
        for selector in ["form.cart", "button[name='add-to-cart']", "input[name='product_id']"]:
            el = soup.select_one(selector)
            if not el:
                continue
            for attr in ["data-product_id", "value"]:
                val = clean_text(el.get(attr, ""))
                if val and val.isdigit():
                    product_id = val
                    break
            if product_id:
                break
        if product_id:
            sku = f"CARTECH-{product_id}"
        else:
            slug = urlparse(product_url).path.rstrip("/").split("/")[-1]
            sku = f"CARTECH-{slug}"
    return sku


def extract_description(soup: BeautifulSoup, product_json: dict) -> str:
    def select_text(selector: str) -> str:
        el = soup.select_one(selector)
        return clean_text(el.get_text(" ", strip=True)) if el else ""

    parts = []
    if product_json.get("description"):
        parts.append(clean_text(str(product_json.get("description"))))
    for selector in [
        ".woocommerce-product-details__short-description",
        "#tab-description",
        ".woocommerce-Tabs-panel--description",
        "div[itemprop='description']",
    ]:
        txt = select_text(selector)
        if txt and txt not in parts:
            parts.append(txt)
    return "\n\n".join(unique_text_keep_order(parts))


def allowed_product_image(url: str) -> bool:
    if not url:
        return False
    l = url.lower()
    blocked = [
        "favicon", "logo", "placeholder", "avatar", "spinner", "/themes/", "/plugins/",
        "facebook", "instagram", "tiktok", "data:image",
    ]
    return not any(b in l for b in blocked)


def extract_product_gallery_images(soup: BeautifulSoup, product_json: dict) -> List[str]:
    """Extract only the product gallery images, not images from the description tab."""
    images: List[str] = []

    # WooCommerce gallery: main image and thumbnails are under .woocommerce-product-gallery.
    for el in soup.select(".woocommerce-product-gallery .woocommerce-product-gallery__image a[href], .woocommerce-product-gallery a[href]"):
        href = el.get("href")
        if href:
            images.append(href)
    for el in soup.select(".woocommerce-product-gallery .woocommerce-product-gallery__image img, .woocommerce-product-gallery img"):
        for attr in ["data-large_image", "data-o_src", "data-src", "data-thumb", "src"]:
            val = el.get(attr)
            if val:
                images.append(val)

    # Fallback only if the gallery selector returned nothing. JSON-LD/OG images should be product images,
    # while description images are not used here.
    if not images:
        img_field = product_json.get("image")
        if isinstance(img_field, str):
            images.append(img_field)
        elif isinstance(img_field, list):
            images.extend([str(x) for x in img_field])
        for meta_sel in ['meta[property="og:image"]', 'meta[name="twitter:image"]']:
            for meta in soup.select(meta_sel):
                if meta.get("content"):
                    images.append(meta["content"])

    return [i for i in unique_keep_order(images) if allowed_product_image(i)][:MAX_IMAGES]


def extract_price(soup: BeautifulSoup, product_json: dict) -> str:
    price = ""
    offers = product_json.get("offers") if isinstance(product_json, dict) else None
    if isinstance(offers, list) and offers:
        offers = offers[0]
    if isinstance(offers, dict):
        currency = clean_text(str(offers.get("priceCurrency", "") or ""))
        offered_price = clean_text(str(offers.get("price", "") or ""))
        if offered_price and (not currency or currency.upper() == "EUR"):
            price = offered_price

    if not price:
        price_html = " "
        for selector in [".summary .price", ".product .summary .price", ".price"]:
            el = soup.select_one(selector)
            if el:
                price_html += " " + el.get_text(" ", strip=True)
        price = parse_price_eur(price_html)

    if not price:
        price = parse_price_eur(soup.get_text(" ", strip=True))

    return price


def parse_variation_price_eur(variation: dict, default_price: str) -> str:
    # Prefer price_html when it contains EUR. If not present, keep the product-level EUR price.
    price_html = clean_text(str(variation.get("price_html", "") or ""))
    price = parse_price_eur(price_html)
    return price or default_price


def parse_product(url: str, category_hint: str = "") -> List[ProductRow]:
    html_text = fetch(url)
    if not html_text:
        return [ProductRow(product_url=url, category=category_hint)]

    soup = BeautifulSoup(html_text, "lxml")
    product_json = first_product_json_ld(soup)

    def select_text(selector: str) -> str:
        el = soup.select_one(selector)
        return clean_text(el.get_text(" ", strip=True)) if el else ""

    name = clean_text(product_json.get("name", ""))
    if not name:
        name = select_text("h1.product_title") or select_text("h1.entry-title")
    if not name:
        og = soup.select_one('meta[property="og:title"]')
        name = clean_text(og.get("content", "")) if og else ""

    base_sku = extract_sku(soup, product_json, url)
    category = choose_category(soup, product_json, category_hint, name, url)
    description = extract_description(soup, product_json)
    images = extract_product_gallery_images(soup, product_json)
    price = extract_price(soup, product_json)

    variations = extract_variations(soup, html_text)
    rows: List[ProductRow] = []
    if variations:
        for variation in variations:
            variation_sku = clean_text(str(variation.get("sku", "") or ""))
            variation_id = clean_text(str(variation.get("variation_id", "") or ""))
            sku = variation_sku or (f"{base_sku}-{variation_id}" if variation_id else base_sku)
            if not sku:
                sku = f"CARTECH-{variation_id or len(rows) + 1}"
            rows.append(ProductRow(
                sku=sku,
                product_name=name,
                category=category,
                description=description,
                images=images,
                price_eur=parse_variation_price_eur(variation, price),
                product_url=url,
            ))
    else:
        rows.append(ProductRow(
            sku=base_sku,
            product_name=name,
            category=category,
            description=description,
            images=images,
            price_eur=price,
            product_url=url,
        ))

    return rows


def dedupe_rows(rows: List[ProductRow]) -> List[ProductRow]:
    seen = set()
    out = []
    for row in rows:
        key = row.sku or row.product_url
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def write_xlsx(products: List[ProductRow], output_file: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "Уникален SKU номер",
        "Име на продукт",
        "Категория",
        "Описание от страницата на продукта",
    ] + [f"Изображение {i}" for i in range(1, MAX_IMAGES + 1)] + [
        "Цена в евро",
        "Линк към продукта",
    ]
    ws.append(headers)

    for p in products:
        imgs = (p.images or [])[:MAX_IMAGES]
        imgs = imgs + [""] * (MAX_IMAGES - len(imgs))
        ws.append([
            p.sku,
            p.product_name,
            p.category,
            p.description,
            *imgs,
            p.price_eur,
            p.product_url,
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        1: 24,
        2: 60,
        3: 34,
        4: 90,
        12: 14,
        13: 60,
    }
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 48)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 70

    wb.save(output_file)


def main() -> None:
    log.info("Starting CarTechBG scraper for %s", BASE_URL)

    # 1) Crawl user-supplied menu categories first, so every product gets a category hint
    # matching the screenshots.
    category_urls, category_hints = collect_category_product_urls()

    # 2) Merge with sitemap to avoid missing products not present in the menu crawl.
    sitemap_urls = collect_sitemap_product_urls()
    urls = sorted(set(category_urls) | set(sitemap_urls))

    if not urls:
        log.warning("No product URLs from categories/sitemap. Falling back to /shop/ crawling.")
        urls = collect_shop_product_urls()

    if len(urls) < 20:
        urls = sorted(set(urls) | set(collect_shop_product_urls()))

    if MAX_PRODUCTS > 0:
        urls = urls[:MAX_PRODUCTS]

    log.info("Total product URLs to scrape: %s", len(urls))
    rows: List[ProductRow] = []
    for idx, url in enumerate(urls, 1):
        log.info("[%s/%s] %s", idx, len(urls), url)
        try:
            hint = category_hints.get(canonical_product_url(url), "")
            rows.extend(parse_product(url, hint))
        except Exception as exc:
            log.exception("Failed parsing %s: %s", url, exc)
            rows.append(ProductRow(product_url=url, category=category_hints.get(canonical_product_url(url), "")))
        time.sleep(REQUEST_SLEEP)

    rows = dedupe_rows(rows)
    write_xlsx(rows, OUTPUT_FILE)
    log.info("Done. Wrote %s rows to %s", len(rows), OUTPUT_FILE)


if __name__ == "__main__":
    main()
